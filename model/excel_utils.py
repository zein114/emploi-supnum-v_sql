"""
Module contenant les fonctions d'export Excel pour la génération des emplois du temps.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import shutil
from datetime import datetime
import mysql.connector
from load_data import DB_CONFIG

def export_timetables_to_single_excel(solver_results, Groupes_Principale, Sous_Groupes, Sous_Group_Id_Map, Sous_Group_Reference_Group, Group_Id_Map, Matieres, ProCM, ProTP, ProTD, J, GP, GT, Matiere_Codes, All_Rooms, output_dir, days=None, time_slots=None):
    """
    Exporte tous les emplois du temps des groupes dans un seul fichier Excel avec plusieurs feuilles.
    """
    X, Y, Z, W, U_TD, U_TP = solver_results
    
    # Définition des jours et des créneaux horaires
    if days is None or len(days) == 0:
        days = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi']
    
    if time_slots is None or len(time_slots) == 0:
        time_slots = ['08:00-9:30', '9:45-11:15', '11:30-13:00', '15:00-16:30', '17:00-18:30']

    num_slots_per_day = len(time_slots)
    K = len(days) * num_slots_per_day
    
    # Reverse mapping: Sub-group (GT index) -> Principal group (GP index)
    GT_to_GP_Map = {}
    for gid, gt_idx in Sous_Group_Id_Map.items():
        if gid in Group_Id_Map:
            GT_to_GP_Map[gt_idx] = Group_Id_Map[gid]

    # Pré-calcul des affectations de salles pour chaque créneau k
    slot_room_assignments = {} # k -> liste de (type_session, g_ou_gt, j, nom_salle)
    
    for k in range(K):
        assigned_rooms_at_k = set()
        slot_room_assignments[k] = []
        
        # 1. Identifier les sessions de CM au créneau k
        for g in range(GP):
            for j in range(J):
                if X[g][j][k].solution_value() > 0.5:
                    # Trouver une salle de CM libre
                    room_found = "N/A"
                    for r in All_Rooms:
                        if r['Type'] == 'CM' and r['Salle'] not in assigned_rooms_at_k:
                            room_found = r['Salle']
                            assigned_rooms_at_k.add(room_found)
                            break
                    slot_room_assignments[k].append(('CM', g, j, room_found))
                
                if W[g][j][k].solution_value() > 0.5:
                    slot_room_assignments[k].append(('CM Online', g, j, 'En ligne'))
                    
                # 1.5 Identifier les sessions de TD au créneau k (Maintenant Principal)
                if Z[g][j][k].solution_value() > 0.5:
                    room_found = "N/A"
                    for r in All_Rooms:
                        if r['Type'] in ['TD', 'CM'] and r['Salle'] not in assigned_rooms_at_k:
                            room_found = r['Salle']
                            assigned_rooms_at_k.add(room_found)
                            break
                    slot_room_assignments[k].append(('TD', g, j, room_found))
                
                # TD Online (Maintenant Principal)
                if U_TD[g][j][k].solution_value() > 0.5:
                    slot_room_assignments[k].append(('TD Online', g, j, 'En ligne'))
        
        # 2. Identifier les sessions de TP/OnlineSub au créneau k
        for gt in range(GT):
            for j in range(J):
                if Y[gt][j][k].solution_value() > 0.5:
                    room_found = "N/A"
                    for r in All_Rooms:
                        if r['Type'] == 'TP' and r['Salle'] not in assigned_rooms_at_k:
                            room_found = r['Salle']
                            assigned_rooms_at_k.add(room_found)
                            break
                    slot_room_assignments[k].append(('TP', gt, j, room_found))

    # Création du nom du fichier de sortie
    output_file = "Tous_les_Emplois_du_Temps.xlsx"
    
    # Records collected for database insertion
    db_records = []
    
    # Création d'un nouveau classeur
    wb = Workbook()
    wb.remove(wb.active)  # Supprimer la feuille par défaut
    
    # Création d'une feuille pour chaque groupe
    for g in range(GP):
        group_name = Groupes_Principale[g]
        
        # Création d'une nouvelle feuille (limite le nom à 31 caractères)
        sheet_name = group_name[:31] if len(group_name) > 31 else group_name
        ws = wb.create_sheet(title=sheet_name)
        
        # Ajout du titre
        last_col_letter = get_column_letter(1 + len(days))
        ws.merge_cells(f'A1:{last_col_letter}1')
        title_cell = ws['A1']
        title_cell.font = Font(size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.value = f"Emploi du Temps - {group_name}"
        ws.row_dimensions[1].height = 30
        
        # Ajout des en-têtes (jours)
        ws['A2'] = "Horaire"
        ws['A2'].font = Font(bold=True)
        ws['A2'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
        
        for col_idx, day_info in enumerate(days, start=2):
            day_name = day_info['name'] if isinstance(day_info, dict) else day_info
            cell = ws.cell(row=2, column=col_idx)
            cell.value = day_name
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Construction des données de l'emploi du temps
        for row_idx, time_slot in enumerate(time_slots, start=3):
            slot_range = time_slot
            slot_is_active = 1
            if isinstance(time_slot, dict):
                slot_range = time_slot.get('time_range', str(time_slot))
                slot_is_active = time_slot.get('is_active', 1)
            
            time_cell = ws.cell(row=row_idx, column=1)
            time_cell.value = slot_range
            time_cell.font = Font(bold=True)
            time_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            time_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            for col_idx, day_info in enumerate(days, start=2):
                day_idx = col_idx - 2
                is_active = day_info.get('is_active', 1) if isinstance(day_info, dict) else 1
                slot_idx = row_idx - 3
                k = day_idx * num_slots_per_day + slot_idx
                
                cell = ws.cell(row=row_idx, column=col_idx)
                session_info = ""
                found = False

                if is_active == 0 or slot_is_active == 0:
                    session_info = "x"
                    found = True
                else:
                    sessions_list = []
                    
                    # CM et CM Online
                    for j in range(J):
                        if X[g][j][k].solution_value() > 0.5:
                            prof = ProCM[j].get(g, "CM") if isinstance(ProCM[j], dict) else (ProCM[j][0] if ProCM[j] else "CM")
                            code = Matiere_Codes[j]
                            room = next((r[3] for r in slot_room_assignments[k] if r[0]=='CM' and r[1]==g and r[2]==j), "N/A")
                            session_str = f"[{code}] {Matieres[j]}\n(CM)\nProf: {prof}\nSalle: {room}"
                            sessions_list.append(session_str)
                            
                        if W[g][j][k].solution_value() > 0.5:
                            prof = ProCM[j].get(g, "CM Online") if isinstance(ProCM[j], dict) else (ProCM[j][0] if ProCM[j] else "CM Online")
                            code = Matiere_Codes[j]
                            session_str = f"[{code}] {Matieres[j]}\n(CM Online)\nProf: {prof}\nSalle: En ligne"
                            sessions_list.append(session_str)
                            
                        # TD (Principal)
                        if Z[g][j][k].solution_value() > 0.5:
                            prof = ProTD[j].get(g, "TD") if isinstance(ProTD[j], dict) else (ProTD[j][0] if ProTD[j] else "TD")
                            code = Matiere_Codes[j]
                            room = next((r[3] for r in slot_room_assignments[k] if r[0]=='TD' and r[1]==g and r[2]==j), "N/A")
                            session_str = f"[{code}] {Matieres[j]}\n(TD)\nProf: {prof}\nSalle: {room}"
                            sessions_list.append(session_str)
                            
                        # TD Online (Principal)
                        if U_TD[g][j][k].solution_value() > 0.5:
                            prof = ProTD[j].get(g, "TD Online") if isinstance(ProTD[j], dict) else (ProTD[j][0] if ProTD[j] else "TD Online")
                            code = Matiere_Codes[j]
                            session_str = f"[{code}] {Matieres[j]}\n(TD Online)\nProf: {prof}\nSalle: En ligne"
                            sessions_list.append(session_str)
                    
                    # Sous-groupes (TD, TP, TD Online, TP Online)
                    group_id_val = None
                    for gid_val, idx in Group_Id_Map.items():
                        if idx == g:
                            group_id_val = gid_val
                            break
                    
                    merge_from_ids = [sid for sid, refs in Sous_Group_Reference_Group.items() if group_id_val in [r.strip() for r in str(refs).split(',')]]
                    
                    subgroups_of_g = []
                    for sid, pref in Sous_Group_Reference_Group.items():
                        # Case 1: Subgroup of the principal group itself
                        if pref == group_id_val:
                            subgroups_of_g.append(sid)
                        # Case 2: Subgroup of a group that merges into this principal group
                        elif pref in merge_from_ids:
                            subgroups_of_g.append(sid)

                    subgroup_indices = [Sous_Group_Id_Map[si] for si in subgroups_of_g if si in Sous_Group_Id_Map]
                    
                    # 'merge_from_ids' contains the IDs of the specialty/languages groups (from groups_principale)
                    linked_gp_indices = [Group_Id_Map[si] for si in merge_from_ids if si in Group_Id_Map and Group_Id_Map[si] != g]
                    
                    for j in range(J):
                        # Sessions of linked principal groups (e.g., Specialty groups)
                        for lgp in linked_gp_indices:
                            # CM
                            if X[lgp][j][k].solution_value() > 0.5:
                                prof = ProCM[j].get(lgp, "CM") if isinstance(ProCM[j], dict) else (ProCM[j][0] if ProCM[j] else "CM")
                                code = Matiere_Codes[j]
                                room = next((r[3] for r in slot_room_assignments[k] if r[0]=='CM' and r[1]==lgp and r[2]==j), "N/A")
                                name = Groupes_Principale[lgp]
                                session_str = f"[{code}] {Matieres[j]}\n(CM) - {name}\nProf: {prof}\nSalle: {room}"
                                sessions_list.append(session_str)
                            
                            # CM Online
                            if W[lgp][j][k].solution_value() > 0.5:
                                prof = ProCM[j].get(lgp, "CM Online") if isinstance(ProCM[j], dict) else (ProCM[j][0] if ProCM[j] else "CM Online")
                                code = Matiere_Codes[j]
                                name = Groupes_Principale[lgp]
                                session_str = f"[{code}] {Matieres[j]}\n(CM Online) - {name}\nProf: {prof}\nSalle: En ligne"
                                sessions_list.append(session_str)
                                
                            # TD
                            if Z[lgp][j][k].solution_value() > 0.5:
                                prof = ProTD[j].get(lgp, "TD") if isinstance(ProTD[j], dict) else (ProTD[j][0] if ProTD[j] else "TD")
                                code = Matiere_Codes[j]
                                room = next((r[3] for r in slot_room_assignments[k] if r[0]=='TD' and r[1]==lgp and r[2]==j), "N/A")
                                name = Groupes_Principale[lgp]
                                session_str = f"[{code}] {Matieres[j]}\n(TD) - {name}\nProf: {prof}\nSalle: {room}"
                                sessions_list.append(session_str)
                                
                            # TD Online
                            if U_TD[lgp][j][k].solution_value() > 0.5:
                                prof = ProTD[j].get(lgp, "TD Online") if isinstance(ProTD[j], dict) else (ProTD[j][0] if ProTD[j] else "TD Online")
                                code = Matiere_Codes[j]
                                name = Groupes_Principale[lgp]
                                session_str = f"[{code}] {Matieres[j]}\n(TD Online) - {name}\nProf: {prof}\nSalle: En ligne"
                                sessions_list.append(session_str)

                        # Check for TP sessions in subgroups (Standard subgroups)
                        active_tp = [si for si in subgroup_indices if Y[si][j][k].solution_value() > 0.5]
                        active_onl_tp = [si for si in subgroup_indices if U_TP[si][j][k].solution_value() > 0.5]
                        
                        # Combiner par type pour l'affichage
                        for sess_type, indices in [("TP", active_tp), ("TP Online", active_onl_tp)]:
                            if not indices:
                                continue
                            
                            base_type = "TP" if "TP" in sess_type else "TD"
                            prof_map = ProTP[j] if base_type == "TP" else ProTD[j]
                            code = Matiere_Codes[j]
                            
                            # Grouper les sous-groupes par professeur
                            prof_groups = {}
                            for si in indices:
                                p = prof_map.get(si, sess_type) if isinstance(prof_map, dict) else (prof_map[0] if prof_map else sess_type)
                                if p not in prof_groups:
                                    prof_groups[p] = []
                                prof_groups[p].append(si)
                                
                            for prof, prof_indices in prof_groups.items():
                                if len(prof_indices) == len(subgroup_indices) and len(subgroup_indices) > 1:
                                    sg_detail = "G-S Complet"
                                else:
                                    # Rename subgroups if it's a TP session
                                    sg_names = []
                                    for si in prof_indices:
                                        name = Sous_Groupes[si]
                                        if "TP" in sess_type:
                                            name = name.replace("-TD", "-TP")
                                        sg_names.append(name)
                                    sg_names = sorted(sg_names)
                                    sg_detail = ", ".join(sg_names)
                                
                                rooms = []
                                for idx_sg in prof_indices:
                                    if "Online" in sess_type:
                                        r = "En ligne"
                                    else:
                                        r = next((r[3] for r in slot_room_assignments[k] if r[0]==sess_type and r[1]==idx_sg and r[2]==j), "N/A")
                                    if r not in rooms: rooms.append(r)
                                room_str = ", ".join(rooms)
                                
                                # Clean professor name: If type is returned as prof, replace with Non assigné
                                display_prof = prof
                                if display_prof in ["TP", "TD", "TP Online", "TD Online"]:
                                    display_prof = "Non assigné"

                                # Clean type display: If group name already shows the type, we can simplify
                                # But let's keep it consistent: always show type in parentheses, but clean group names
                                # Actually, the user wants DSI2-TP2 instead of DSI2-TP2(TP)
                                session_str = f"[{code}] {Matieres[j]}\n({sess_type}) - {sg_detail}\nProf: {display_prof}\nSalle: {room_str}"
                                
                                # If it's a subgroup and sess_type is TP/TD, and sg_detail has the type, 
                                # we could hide (sess_type). But usually the format is preserved for consistency.
                                # However, to match the user's "remove parentheses" request:
                                if sg_detail == group_name or ("-TP" in sg_detail or "-TD" in sg_detail):
                                     # Just ensure no redundant label if it looks like DSI2-TP2(TP )
                                     pass

                                sessions_list.append(session_str)


                    if sessions_list:
                        # Trier les sessions pour s'assurer que TD1/TP1 apparaissent avant TD2/TP2
                        def session_sort_key(s):
                            if "(CM)" in s: return (0, "")
                            if "G-S Complet" in s: return (1, "0")
                            # Extraire les noms des sous-groupes de la chaîne de session pour les utiliser comme clé de tri
                            # Le format est : ...\n(TP/TD) - TD1, TD2\n...
                            try:
                                if " - " in s:
                                    sub_part = s.split(" - ")[1].split("\n")[0]
                                    return (1, sub_part)
                            except:
                                pass
                            return (1, "z")
                        
                        sessions_list.sort(key=session_sort_key)
                        session_info = " /// ".join(sessions_list)
                    else:
                        session_info = ""

                cell.value = session_info if session_info else ""
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                # Enregistrer dans la base de données
                if cell.value and cell.value != "x":
                    day_name = day_info['name'] if isinstance(day_info, dict) else day_info
                    slot_range_db = time_slot.get('time_range', str(time_slot)) if isinstance(time_slot, dict) else time_slot
                    db_records.append((group_name, day_name, slot_range_db, cell.value))
                
                # Colorer la cellule selon le type de session
                if cell.value:
                    if "(CM)" in cell.value:
                        cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
                    elif "(TP)" in cell.value:
                        cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
                    elif "(TD)" in cell.value:
                        cell.fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
                
                # Ajouter des bordures
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.border = thin_border
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 15
        for col in range(2, 2 + len(days)):
            ws.column_dimensions[get_column_letter(col)].width = 30
        
        # Ajuster la hauteur des lignes
        for row in range(3, 3 + len(time_slots)):
            ws.row_dimensions[row].height = 80
    
    # Enregistrer le classeur
    archives_dir = output_dir + "archives_timetables"   
    if not os.path.exists(archives_dir):
        os.makedirs(archives_dir)
    # Archiver l'ancien fichier s'il existe
    if os.path.exists(output_dir + output_file):
        os.replace(output_dir + output_file, archives_dir + "/Tous_les_Emplois_du_Temps_" + datetime.now().strftime("%Y%m%d_%H%M%S") + ".xlsx")
    
    # Enregistrer le nouveau classeur
    wb.save(output_dir + output_file)
    print(f"Tous les emplois du temps ont été exportés vers {output_dir + output_file}")
    
    # Enregistrer dans la base de données
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor()
        cursor.execute("TRUNCATE TABLE `timetables`")
        if db_records:
            insert_query = "INSERT INTO `timetables` (`group_name`, `day`, `time_slot`, `session_info`) VALUES (%s, %s, %s, %s)"
            cursor.executemany(insert_query, db_records)
            conn.commit()
        conn.close()
        print("Les emplois du temps ont été sauvegardés dans la base de données avec succès.")
    except Exception as e:
        print(f"Erreur lors de la sauvegarde dans la base de données: {e}")

