"""
Module contenant les fonctions d'export Excel pour la génération des emplois du temps.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import shutil
from datetime import datetime

def export_timetables_to_single_excel(solver_results, Groupes_Principale, Sous_Groupes, Sous_Group_Code_Map, Sous_Groupes_Reference_Groupes, Group_Code_Map, Matieres, ProCM, ProTP, ProTD, J, GP, GT, Matiere_Codes, All_Rooms, output_dir, days=None, time_slots=None, Pon=None, Son_td=None, Son_tp=None):
    """
    Exporte tous les emplois du temps des groupes dans un seul fichier Excel avec plusieurs feuilles.
    """
    X, Y, Z = solver_results
    
    # Définition des jours et des créneaux horaires
    if days is None or len(days) == 0:
        days = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi']
    
    if time_slots is None or len(time_slots) == 0:
        time_slots = ['08:00-9:30', '9:45-11:15', '11:30-13:00', '15:00-16:30', '17:00-18:30']

    num_slots_per_day = len(time_slots)
    K = len(days) * num_slots_per_day
    
    # Pré-calcul des affectations de salles pour chaque créneau k
    slot_room_assignments = {} # k -> liste de (type_session, g_ou_gt, j, nom_salle)
    
    for k in range(K):
        assigned_rooms_at_k = set()
        slot_room_assignments[k] = []
        
        # 1. Identifier les sessions de CM au créneau k
        for g in range(GP):
            for j in range(J):
                if X[g][j][k].solution_value() > 0.5:
                    # Check if online (Pon > 0)
                    is_online = (Pon is not None and j < len(Pon) and g < len(Pon[j]) and Pon[j][g] > 0)
                    
                    if is_online:
                        slot_room_assignments[k].append(('CM', g, j, 'En ligne'))
                    else:
                        # Trouver une salle de CM libre
                        room_found = "N/A"
                        for r in All_Rooms:
                            if r['Type'] == 'CM' and r['Salle'] not in assigned_rooms_at_k:
                                room_found = r['Salle']
                                assigned_rooms_at_k.add(room_found)
                                break
                        slot_room_assignments[k].append(('CM', g, j, room_found))
        
        # 2. Identifier les sessions de TP/TD au créneau k
        for gt in range(GT):
            for j in range(J):
                # TP session
                if Y[gt][j][k].solution_value() > 0.5:
                    is_online_tp = (Son_tp is not None and j < len(Son_tp) and gt < len(Son_tp[j]) and Son_tp[j][gt] > 0)
                    
                    if is_online_tp:
                        slot_room_assignments[k].append(('TP', gt, j, 'En ligne'))
                    else:
                        room_found = "N/A"
                        for r in All_Rooms:
                            if r['Type'] == 'TP' and r['Salle'] not in assigned_rooms_at_k:
                                room_found = r['Salle']
                                assigned_rooms_at_k.add(room_found)
                                break
                        slot_room_assignments[k].append(('TP', gt, j, room_found))
                
                # TD session
                if Z[gt][j][k].solution_value() > 0.5:
                    is_online_td = (Son_td is not None and j < len(Son_td) and gt < len(Son_td[j]) and Son_td[j][gt] > 0)
                    
                    if is_online_td:
                        slot_room_assignments[k].append(('TD', gt, j, 'En ligne'))
                    else:
                        room_found = "N/A"
                        for r in All_Rooms:
                            if r['Type'] in ['TP', 'TD', 'CM'] and r['Salle'] not in assigned_rooms_at_k:
                                room_found = r['Salle']
                                assigned_rooms_at_k.add(room_found)
                                break
                        slot_room_assignments[k].append(('TD', gt, j, room_found))

    # Création du nom du fichier de sortie
    output_file = "Tous_les_Emplois_du_Temps.xlsx"
    
    # Création d'un nouveau classeur
    wb = Workbook()
    wb.remove(wb.active)  # Supprimer la feuille par défaut
    
    # Création d'une feuille pour chaque groupe
    for g in range(GP):
        group_name = Groupes_Principale[g]
        
        # Création d'une nouvelle feuille (limite le nom à 31 caractères)
        sheet_name = group_name[:31] if len(group_name) > 31 else group_name
        ws = wb.create_sheet(title=sheet_name)
        
        # Ajout du titre
        last_col_letter = get_column_letter(1 + len(days))
        ws.merge_cells(f'A1:{last_col_letter}1')
        title_cell = ws['A1']
        title_cell.value = f"Emploi du Temps - {group_name}"
        title_cell.font = Font(size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        
        # Ajout des en-têtes (jours)
        ws['A2'] = "Horaire"
        ws['A2'].font = Font(bold=True)
        ws['A2'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
        
        for col_idx, day_info in enumerate(days, start=2):
            day_name = day_info['name'] if isinstance(day_info, dict) else day_info
            cell = ws.cell(row=2, column=col_idx)
            cell.value = day_name
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Construction des données de l'emploi du temps
        for row_idx, time_slot in enumerate(time_slots, start=3):
            # Colonne des créneaux horaires
            slot_range = time_slot
            slot_is_active = 1
            if isinstance(time_slot, dict):
                slot_range = time_slot.get('time_range', str(time_slot))
                slot_is_active = time_slot.get('is_active', 1)
            
            time_cell = ws.cell(row=row_idx, column=1)
            time_cell.value = slot_range
            time_cell.font = Font(bold=True)
            time_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            time_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Remplissage pour chaque jour
            for col_idx, day_info in enumerate(days, start=2):
                day_idx = col_idx - 2
                is_active = day_info.get('is_active', 1) if isinstance(day_info, dict) else 1
                slot_idx = row_idx - 3
                k = day_idx * num_slots_per_day + slot_idx
                
                cell = ws.cell(row=row_idx, column=col_idx)
                session_info = ""
                found = False

                if is_active == 0 or slot_is_active == 0:
                    session_info = "x"
                    found = True
                else:
                    sessions_list = []
                    
                    # CM et CM Online
                    for j in range(J):
                        if X[g][j][k].solution_value() > 0.5:
                            # Check if online (Pon > 0)
                            is_online = (Pon is not None and j < len(Pon) and g < len(Pon[j]) and Pon[j][g] > 0)
                            
                            prof = ProCM[j][0] if ProCM[j] else "CM"
                            code = Matiere_Codes[j]
                            
                            if is_online:
                                session_str = f"[{code}] {Matieres[j]}\n(CM Online)\nProf: {prof}\nSalle: En ligne"
                            else:
                                room = next((r[3] for r in slot_room_assignments[k] if r[0]=='CM' and r[1]==g and r[2]==j), "N/A")
                                session_str = f"[{code}] {Matieres[j]}\n(CM)\nProf: {prof}\nSalle: {room}"
                            
                            sessions_list.append(session_str)
                    
                    # Sous-groupes (TD, TP, TD Online, TP Online)
                    group_code_val = None
                    for code_val, idx in Group_Code_Map.items():
                        if idx == g:
                            group_code_val = code_val
                            break
                    
                    subgroups_of_g = []
                    for sub_code, ref_val in Sous_Groupes_Reference_Groupes.items():
                        refs = [r.strip() for r in str(ref_val).split(',')]
                        if group_code_val in refs:
                            subgroups_of_g.append(sub_code)
                    
                    subgroup_indices = [Sous_Group_Code_Map[sc] for sc in subgroups_of_g if sc in Sous_Group_Code_Map]
                    
                    for j in range(J):
                        # Detect all active types including online
                        active_tp = [si for si in subgroup_indices if Y[si][j][k].solution_value() > 0.5 and (Son_tp is None or j >= len(Son_tp) or si >= len(Son_tp[j]) or Son_tp[j][si] == 0)]
                        active_td = [si for si in subgroup_indices if Z[si][j][k].solution_value() > 0.5 and (Son_td is None or j >= len(Son_td) or si >= len(Son_td[j]) or Son_td[j][si] == 0)]
                        active_onl_tp = [si for si in subgroup_indices if Son_tp is not None and j < len(Son_tp) and si < len(Son_tp[j]) and Son_tp[j][si] > 0 and Y[si][j][k].solution_value() > 0.5]
                        active_onl_td = [si for si in subgroup_indices if Son_td is not None and j < len(Son_td) and si < len(Son_td[j]) and Son_td[j][si] > 0 and Z[si][j][k].solution_value() > 0.5]
                        
                        # Process each category
                        for sess_label, indices in [("TP", active_tp), ("TD", active_td), ("TP Online", active_onl_tp), ("TD Online", active_onl_td)]:
                            if indices:
                                base_type = "TP" if "TP" in sess_label else "TD"
                                prof_list = ProTP[j] if base_type == "TP" else ProTD[j]
                                prof = prof_list[0] if prof_list else base_type
                                code = Matiere_Codes[j]
                                
                                if len(indices) == len(subgroup_indices) and len(subgroup_indices) > 1:
                                    sg_detail = "G-S Complet"
                                else:
                                    sg_names = sorted([Sous_Groupes[si] for si in indices])
                                    sg_detail = ", ".join(sg_names)
                                
                                # Gather rooms
                                rooms = []
                                for idx_sg in indices:
                                    if "Online" in sess_label:
                                        r = "En ligne"
                                    else:
                                        r = next((r[3] for r in slot_room_assignments[k] if r[0]==base_type and r[1]==idx_sg and r[2]==j), "N/A")
                                    if r not in rooms: rooms.append(r)
                                room_str = ", ".join(rooms)
                                
                                session_str = f"[{code}] {Matieres[j]}\n({sess_label}) - {sg_detail}\nProf: {prof}\nSalle: {room_str}"
                                sessions_list.append(session_str)

                    if sessions_list:
                        # Trier les sessions pour s'assurer que TD1/TP1 apparaissent avant TD2/TP2
                        def session_sort_key(s):
                            if "(CM)" in s: return (0, "")
                            if "G-S Complet" in s: return (1, "0")
                            # Extraire les noms des sous-groupes de la chaîne de session pour les utiliser comme clé de tri
                            try:
                                if " - " in s:
                                    sub_part = s.split(" - ")[1].split("\n")[0]
                                    return (1, sub_part)
                            except:
                                pass
                            return (1, "z")
                        
                        sessions_list.sort(key=session_sort_key)
                        session_info = " /// ".join(sessions_list)
                    else:
                        session_info = ""

                cell.value = session_info if session_info else ""
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                # Colorer la cellule selon le type de session
                if cell.value:
                    if "Online" in cell.value:
                        # Light blue for online
                        cell.fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
                    elif "(CM)" in cell.value:
                        cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
                    elif "(TP)" in cell.value:
                        cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
                    elif "(TD)" in cell.value:
                        cell.fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
                
                # Ajouter des bordures
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.border = thin_border
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 15
        for col in range(2, 2 + len(days)):
            ws.column_dimensions[get_column_letter(col)].width = 30
        
        # Ajuster la hauteur des lignes
        for row in range(3, 3 + len(time_slots)):
            ws.row_dimensions[row].height = 80
    
    # Enregistrer le classeur
    archives_dir = output_dir + "archives_timetables"   
    if not os.path.exists(archives_dir):
        os.makedirs(archives_dir)
    # Archiver l'ancien fichier s'il existe
    if os.path.exists(output_dir + output_file):
        os.replace(output_dir + output_file, archives_dir + "/Tous_les_Emplois_du_Temps_" + datetime.now().strftime("%Y%m%d_%H%M%S") + ".xlsx")
    
    # Enregistrer le nouveau classeur
    wb.save(output_dir + output_file)
    print(f"Tous les emplois du temps ont été exportés vers {output_dir + output_file}")

